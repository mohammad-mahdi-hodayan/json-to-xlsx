import openpyxl 
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import json

wb = Workbook(write_only = True)
ws = wb.create_sheet()

with open('json_file_path.json', 'r', encoding="utf8") as json_file:
    data = json.load(json_file)
    for a_data in data :
        for main_data in data[a_data]:
            '''main_data["title"]'''
            your_2nd_col = str(main_data["data_name"])
            ws.append([main_data["data_name_col_1"],your_2nd_col)
wb.save('output_file_name.xlsx')
